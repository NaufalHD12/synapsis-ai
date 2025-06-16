# ==============================================================================
# Synapsis AI - Backend Flask
# Author: Naufal Hadi Darmawan
# ==============================================================================

# --- Impor Pustaka ---
import os
import json
import io
import pandas as pd
import requests
from dotenv import load_dotenv
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl.styles import Font, Alignment

# --- Inisialisasi & Konfigurasi ---
load_dotenv()
app = Flask(__name__)

# Konfigurasi dari Environment Variables
DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY")
DEEPSEEK_API_URL = "https://api.deepseek.com/chat/completions"

# --- Template Prompt untuk AI ---
AI_PROMPT_TEMPLATE = """
Anda adalah seorang asisten HR virtual multilingual yang sangat ahli dalam menyaring CV.

TUGAS ANDA:
1. Deteksi bahasa utama yang digunakan dalam CV (misalnya: "Indonesia" atau "English").
2. Analisis CV kandidat berdasarkan Deskripsi Pekerjaan (JD).
3. Berikan respons HANYA dalam format JSON yang valid, TANPA teks tambahan di luar JSON. Gunakan bahasa yang sama dengan yang Anda deteksi di CV untuk mengisi nilai-nilai dalam JSON.

--- DESKRIPSI PEKERJAAN ---
{jd_text}
---

--- CV KANDIDAT ---
{cv_text}
---

Berikan analisis dalam format JSON berikut:
{{
    "nama_kandidat": "Ekstrak nama lengkap kandidat dari CV",
    "skor_kecocokan": <Angka 0-100 yang merepresentasikan kecocokan CV dengan JD>,
    "bahasa_terdeteksi": "Bahasa yang Anda deteksi dari CV (contoh: 'Indonesia' atau 'English')",
    "ringkasan_positif": "Jelaskan dalam 2-3 kalimat mengapa kandidat ini berpotensi cocok, berdasarkan kesesuaian pengalaman dan skill utama.",
    "poin_kunci_cocok": [
        "Sebutkan skill atau pengalaman spesifik pertama dari CV yang paling cocok dengan JD.",
        "Sebutkan skill atau pengalaman spesifik kedua dari CV yang cocok dengan JD.",
        "Sebutkan poin relevan ketiga jika ada."
    ],
    "poin_perhatian": [
        "Sebutkan kualifikasi pertama dari JD yang tidak ditemukan atau kurang cocok di CV.",
        "Sebutkan kualifikasi kedua dari JD yang tidak ditemukan atau kurang cocok di CV."
    ]
}}
"""

# --- Fungsi Helper ---

def call_deepseek_api(jd_text, cv_text):
    """
    Memanggil DeepSeek API dan mengembalikan hasil analisis.
    Menampilkan error jika tidak ditemukan.
    """
    if not DEEPSEEK_API_KEY:
        print("PERINGATAN: API_KEY tidak ditemukan.")

    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {DEEPSEEK_API_KEY}"}
    prompt = AI_PROMPT_TEMPLATE.format(jd_text=jd_text, cv_text=cv_text)
    data = {"model": "deepseek-chat", "messages": [{"role": "system", "content": "You are a multilingual HR assistant that only outputs valid JSON."}, {"role": "user", "content": prompt}], "response_format": {"type": "json_object"}}
    
    try:
        response = requests.post(DEEPSEEK_API_URL, headers=headers, json=data, timeout=180)
        response.raise_for_status()
        return json.loads(response.json()['choices'][0]['message']['content'])
    except Exception as e:
        print(f"ERROR: API call failed: {e}")
        return {"error": f"API call failed: {e}"}

# --- Rute & Endpoint API ---

@app.route('/')
def index():
    """Menampilkan halaman utama frontend."""
    return render_template('index.html')

@app.route('/api/analyze', methods=['POST'])
def analyze_cvs_endpoint():
    """Menerima data JD & CV, lalu mengembalikan hasil analisis dari AI."""
    data = request.get_json()
    if not data or 'jdText' not in data or 'cvData' not in data:
        return jsonify({"error": "Permintaan tidak valid, data tidak lengkap."}), 400

    jd_text, cv_data = data['jdText'], data['cvData']
    results = []
    
    for cv in cv_data:
        try:
            analysis_result = call_deepseek_api(jd_text, cv['text'])
            if "error" in analysis_result:
                analysis_result.update({'nama_kandidat': cv['filename'], 'skor_kecocokan': 0})
            analysis_result['original_filename'] = cv['filename']
            results.append(analysis_result)
        except Exception as e:
            print(f"ERROR: Gagal memproses CV {cv['filename']}: {e}")
            results.append({"nama_kandidat": cv['filename'], "original_filename": cv['filename'], "skor_kecocokan": 0, "error": True, "ringkasan_positif": f"Terjadi kesalahan internal: {e}", "poin_kunci_cocok": [], "poin_perhatian": []})
    
    sorted_results = sorted(results, key=lambda x: x.get('skor_kecocokan', 0), reverse=True)
    return jsonify(sorted_results)

@app.route('/api/download_excel', methods=['POST'])
def download_excel_endpoint():
    """Menerima data JSON dan mengubahnya menjadi laporan Excel yang diformat."""
    results_data = request.get_json()
    if not results_data:
        return jsonify({"error": "Tidak ada data untuk diekspor."}), 400

    df = pd.DataFrame(results_data)
    
    def format_list_to_bullets(lst):
        return "\n".join([f"• {item}" for item in lst]) if isinstance(lst, list) else lst

    df['poin_kunci_cocok'] = df['poin_kunci_cocok'].apply(format_list_to_bullets)
    df['poin_perhatian'] = df['poin_perhatian'].apply(format_list_to_bullets)

    columns_order = ['nama_kandidat', 'skor_kecocokan', 'ringkasan_positif', 'poin_kunci_cocok', 'poin_perhatian', 'bahasa_terdeteksi', 'original_filename']
    df = df.reindex(columns=columns_order)
    
    df.rename(columns={
        'nama_kandidat': 'Nama Kandidat', 'skor_kecocokan': 'Skor', 'ringkasan_positif': 'Ringkasan Positif',
        'poin_kunci_cocok': 'Poin Kunci Cocok', 'poin_perhatian': 'Poin Perhatian',
        'bahasa_terdeteksi': 'Bahasa CV', 'original_filename': 'Nama File Asli'
    }, inplace=True)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Hasil Analisis')
        worksheet = writer.sheets['Hasil Analisis']

        # Definisikan style Font dan Alignment
        bold_font = Font(bold=True)
        wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

        # Terapkan font tebal ke header dan atur lebar kolom
        for i, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=i)
            cell.font = bold_font
        
        column_widths = {'A': 30, 'B': 10, 'C': 50, 'D': 50, 'E': 50, 'F': 15, 'G': 30}
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Terapkan text wrapping untuk semua sel kecuali header
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = wrap_alignment

    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='hasil_penyaringan_synapsis_ai.xlsx'
    )

# --- Main Execution ---
if __name__ == '__main__':
    app.run(debug=True)
